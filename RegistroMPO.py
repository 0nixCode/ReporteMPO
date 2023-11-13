import tkinter as tk
from tkinter import ttk, messagebox, filedialog, Menu
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from PIL import Image, ImageTk
from ttkthemes import ThemedStyle
import os
import sys

version = "5.8 Beta"
data_list = []
required_entries = []
success_label = None

font_size = 9
common_font = ('Helvetica', font_size)
default_column_widths = {'ID': 40, 'Nº Expediente': 120, 'Nombre Comercial': 120, 'Razón Social': 120, 'Celular': 80,
                        'Dirección': 150, 'Referencias': 150}

color_verde = '#4CAF50'
fuente_profesional = ('Helvetica', 10)

# Variable global para almacenar la posición del cursor
last_cursor_position = 0

def format_cellphone_number(cellphone):
    formatted_cellphone = ''.join(cellphone.split())
    formatted_cellphone = " ".join([formatted_cellphone[i:i+3] for i in range(0, len(formatted_cellphone), 3)])
    return formatted_cellphone

def animate_button(button):
    button.state(['pressed'])
    window.after(100, lambda: button.state(['!pressed']))

def is_valid_number(input_string):
    return input_string.isdigit()

def capitalize_first_letter(text):
    return ' '.join(word.capitalize() for word in text.split())

def validate_entry(input_value, entry):
    if not is_valid_number(input_value):
        entry.delete(0, 'end')

def show_warning(message):
    messagebox.showwarning("Advertencia", message)

def update_display():
    global required_entries
    global success_label
    global last_cursor_position

    if not required_entries:
        required_entries = [numero_expediente_entry, nombre_comercial_entry, razon_social_entry, numero_celular_entry,
                            direccion_entry, referencias_entry]

    if not all(entry.get() for entry in required_entries):
        show_warning("Por favor, complete todos los campos antes de guardar.")
        return

    formatted_entries = [capitalize_first_letter(entry.get()) for entry in required_entries]

    numero_expediente, nombre_comercial, razon_social, numero_celular, direccion, referencias = formatted_entries

    numero_celular = numero_celular.replace(" ", "")
    if not is_valid_number(numero_celular) or len(numero_celular) != 9:
        show_warning("Número de celular no es válido.")
        return

    formatted_cellphone = format_cellphone_number(numero_celular)

    data = [
        len(data_list) + 1,
        numero_expediente,
        nombre_comercial,
        razon_social,
        formatted_cellphone, 
        direccion,
        referencias
    ]

    data_list.append(data)

    tree.delete(*tree.get_children())

    for i, row in enumerate(data_list, start=1):
        tree.insert('', 'end', values=[i] + row[1:])

    adjust_column_widths()
    update_status_bar("Datos actualizados exitosamente.", success_label, success=True, bold=True)

    for entry in required_entries:
        entry.delete(0, 'end')

    # Restaurar la posición del cursor después de actualizar el número de teléfono
    numero_celular_entry.icursor(last_cursor_position)
    numero_celular_entry.select_clear()

def clear_fields():
    for entry in required_entries:
        entry.delete(0, 'end')

def export_to_excel():
    global required_entries
    global success_label

    if not required_entries:
        required_entries = [numero_expediente_entry, nombre_comercial_entry, razon_social_entry, numero_celular_entry,
                            direccion_entry, referencias_entry]

    if not data_list:
        show_warning("No hay datos para exportar.")
        return

    try:
        default_filename = "Inspecciones.xlsx"
        default_path = "~"
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos Excel", "*.xlsx")],
                                                  initialfile=default_filename, initialdir=default_path)

        if file_path:
            workbook = Workbook()
            sheet = workbook.active

            header_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center')

            headers = ["ID", "Nº Expediente", "Nombre Comercial", "Razón Social", "Celular", "Dirección", "Referencias"]
            sheet.append(headers)

            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            for data in data_list:
                # Formatear el número de celular antes de agregarlo a la hoja de cálculo
                data_to_export = data.copy()
                data_to_export[4] = format_cellphone_number(data[4])
                sheet.append(data_to_export)

            data_alignment = Alignment(horizontal='center')

            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.alignment = data_alignment

            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column].width = adjusted_width

            workbook.save(file_path)
            messagebox.showinfo("Éxito", f"Datos exportados a {file_path}")

            for entry in required_entries:
                entry.delete(0, 'end')

            adjust_column_widths()
            update_status_bar("Datos exportados exitosamente.", success_label, success=True, bold=True)
    except Exception as e:
        messagebox.showerror("Error", f"Error al exportar a Excel: {e}")
        update_status_bar("Error al exportar a Excel.", success_label, success=False, color="red", bold=True)

def edit_selected_data():
    selected_item = tree.selection()
    if not selected_item:
        return

    index = tree.index(selected_item)

    if not all(entry.get() for entry in required_entries):
        show_warning("Por favor, complete todos los campos antes de editar.")
        return

    formatted_entries = [capitalize_first_letter(entry.get()) for entry in required_entries]

    numero_expediente, nombre_comercial, razon_social, numero_celular, direccion, referencias = formatted_entries

    numero_celular = numero_celular.replace(" ", "")
    if not is_valid_number(numero_celular) or len(numero_celular) != 9:
        show_warning("Número de celular no es válido.")
        return

    # Aquí se vuelve a aplicar el formato al número de celular antes de asignarlo a la lista
    formatted_cellphone = format_cellphone_number(numero_celular)

    data_list[index][1] = numero_expediente
    data_list[index][2] = nombre_comercial
    data_list[index][3] = razon_social
    data_list[index][4] = formatted_cellphone  # Aquí se utiliza el número de celular formateado
    data_list[index][5] = direccion
    data_list[index][6] = referencias

    for entry in required_entries:
        entry.delete(0, 'end')

    tree.delete(*tree.get_children())
    for i, row in enumerate(data_list, start=1):
        tree.insert('', 'end', values=[i] + row[1:])

    adjust_column_widths()
    update_status_bar("Datos editados exitosamente.", success_label, success=True, bold=True)

def delete_selected_data():
    selected_item = tree.selection()
    if not selected_item:
        return
    response = messagebox.askyesno("Eliminar Datos", "¿Estás seguro de que quieres eliminar estos datos?")
    if response:
        index = tree.index(selected_item)
        data_list.pop(index)

        for i, item in enumerate(data_list, start=1):
            item[0] = i

        tree.delete(*tree.get_children())
        for row in data_list:
            tree.insert('', 'end', values=row)

        adjust_column_widths()
        update_status_bar("Datos eliminados exitosamente.", success_label, success=True, bold=True)

def load_selected_data(event):
    selected_item = tree.selection()
    if not selected_item:
        return

    selected_data = tree.item(selected_item, 'values')

    numero_expediente_entry.delete(0, 'end')
    nombre_comercial_entry.delete(0, 'end')
    razon_social_entry.delete(0, 'end')
    numero_celular_entry.delete(0, 'end')
    direccion_entry.delete(0, 'end')
    referencias_entry.delete(0, 'end')

    numero_expediente_entry.insert(0, selected_data[1])
    nombre_comercial_entry.insert(0, selected_data[2])
    razon_social_entry.insert(0, selected_data[3])
    numero_celular_entry.insert(0, selected_data[4])
    direccion_entry.insert(0, selected_data[5])
    referencias_entry.insert(0, selected_data[6])

    adjust_column_widths()

def load_resized_icon(file_path, width, height):
    original_image = Image.open(file_path)
    resampling_filter = Image.ANTIALIAS if hasattr(Image, 'ANTIALIAS') else Image.LANCZOS
    resized_image = original_image.resize((width, height), resampling_filter)
    return ImageTk.PhotoImage(resized_image)

def validate_and_format_phone_number(entry):
    global last_cursor_position
    
    # Obtener la posición del cursor antes de realizar cambios
    last_cursor_position = entry.index(tk.INSERT)
    
    # Obtener el texto actual del campo de entrada
    text = entry.get().replace(" ", "")
    
    # Solo mantener los dígitos (esto valida la entrada)
    digits_only = ''.join(filter(str.isdigit, text))
    
    # Limitar la longitud a 9 dígitos
    digits_only = digits_only[:9]

    # Formatear el número con espacios cada tres dígitos
    formatted_number = " ".join(digits_only[i:i+3] for i in range(0, len(digits_only), 3))

    # Actualizar el campo de entrada con el nuevo formato
    if entry.get() != formatted_number:
        entry.delete(0, tk.END)
        entry.insert(0, formatted_number)
    
    # Restaurar la posición del cursor después de actualizar el número de teléfono
    entry.icursor(last_cursor_position)
    entry.select_clear()

def resource_path(relative_path):
    """ Devuelve la ruta de acceso al recurso, funciona para el desarrollo y para el ejecutable único."""
    if getattr(sys, 'frozen', False):
        # Si estamos en el contexto de PyInstaller, usa el directorio temporal
        base_path = sys._MEIPASS
    else:
        # De lo contrario, usa el directorio de trabajo actual (para el desarrollo)
        base_path = os.path.dirname(os.path.abspath(__file__))
    
    return os.path.join(base_path, relative_path)

def update_status_bar(message, label, success=True, bold=False):
    if success:
        label.config(text=message, foreground="green", font=(common_font[0], font_size, "bold" if bold else "normal"))
    else:
        label.config(text=message, foreground="red", font=(common_font[0], font_size, "bold" if bold else "normal"))

def adjust_column_widths():
    scale_factor = 1.5

    for col in tree["columns"]:
        if col == "ID":
            tree.column(col, width=20, anchor="center")
            continue

        values = [str(tree.set(item, col)) for item in tree.get_children()]
        values = [val for val in values if val]

        if values:
            max_width = max(len(str(value)) for value in values)
            header_width = len(col)
            max_width = max(max_width, header_width)
            adjusted_width = max_width * scale_factor

            tree.column(col, width=int(round(adjusted_width)))

window = tk.Tk()
window.title(f"Registro de expedientes Defensa Civil - Versión {version}")
window.resizable(False, False)
# Ahora establece el icono usando la función resource_path
window.iconbitmap(resource_path('imagenes/defensa.ico'))


style = ThemedStyle(window)
style.set_theme("breeze")

frame = ttk.Frame(window, padding=(20, 20))
frame.grid(row=0, column=0, sticky='nsew', columnspan=3)

user_info_frame = ttk.LabelFrame(frame, text="Datos de expediente", padding=(10, 10))
user_info_frame.grid(row=1, column=0, padx=20, pady=20, sticky='nsew')

labels = ["Nº Expediente", "Nombre Comercial", "Razon Social", "Celular", "Direccion", "Referencias"]
entries = []

for i, label in enumerate(labels):
    lbl = ttk.Label(user_info_frame, text=label, font=common_font)
    lbl.grid(row=i, column=0, pady=(10, 5), padx=(50, 5), sticky='e')

    entry = ttk.Entry(user_info_frame, font=common_font)
    entry.grid(row=i, column=1, pady=(10, 5), padx=(0, 10), sticky='w')
    entries.append(entry)

numero_celular_entry = entries[3]
numero_celular_entry.bind("<KeyRelease>", lambda event: validate_and_format_phone_number(numero_celular_entry))

numero_expediente_entry = entries[0]
numero_expediente_entry.bind("<KeyRelease>", lambda e: validate_entry(numero_expediente_entry.get(), numero_expediente_entry))

nombre_comercial_entry = entries[1]
razon_social_entry = entries[2]
numero_celular_entry = entries[3]
direccion_entry = entries[4]
referencias_entry = entries[5]

columns = ["ID", "Nº Expediente", "Nombre Comercial", "Razón Social", "Celular", "Dirección", "Referencias"]
tree = ttk.Treeview(frame, columns=columns, show="headings")

for col in columns:
    tree.heading(col, text=col)

tree.column("ID", width=20, anchor="center")
tree.column("Nº Expediente", width=120, anchor="center")
tree.column("Nombre Comercial", width=120, anchor="center")
tree.column("Razón Social", width=120, anchor="center")

for col in columns[4:]:
    tree.column(col, width=150, anchor="center")

tree.grid(row=1, column=1, padx=20, pady=20, sticky='nsew')

scrollbar = ttk.Scrollbar(frame, command=tree.yview)
scrollbar.grid(row=1, column=2, sticky='ns')

tree.config(yscrollcommand=scrollbar.set)

context_menu = Menu(window, tearoff=0)
context_menu.add_command(label="Eliminar", command=delete_selected_data)

tree.bind("<Button-3>", lambda event: context_menu.post(event.x_root, event.y_root))
tree.bind("<ButtonRelease-1>", load_selected_data)

button_alimentar_icon = load_resized_icon(resource_path("imagenes/guardar.png"), 25, 25)
button_alimentar = ttk.Button(user_info_frame, image=button_alimentar_icon, text="Guardar", compound="left", command=lambda: [animate_button(button_alimentar), update_display()], style="Button.TButton")
button_alimentar.grid(row=len(labels), column=0, pady=(20, 0), padx=(10, 5), sticky='we')

button_exportar_icon = load_resized_icon(resource_path("imagenes/excel.png"), 25, 25)
button_exportar = ttk.Button(user_info_frame, image=button_exportar_icon, text="Exportar a Excel", compound="left", command=lambda: [animate_button(button_exportar), export_to_excel()], style="Button.TButton")
button_exportar.grid(row=len(labels) + 1, column=0, pady=(5, 0), padx=(10, 5), sticky='we')

button_editar_icon = load_resized_icon(resource_path("imagenes/editar.png"), 25, 25)
button_editar = ttk.Button(user_info_frame, image=button_editar_icon, text="Editar", compound="left", command=lambda: [animate_button(button_editar), edit_selected_data()], style="Button.TButton")
button_editar.grid(row=len(labels), column=1, pady=(20, 0), padx=(5, 10), sticky='we')

button_limpiar_icon = load_resized_icon(resource_path("imagenes/limpiar.png"), 25, 25)
button_limpiar = ttk.Button(user_info_frame, image=button_limpiar_icon, text="Limpiar", compound="left", command=lambda: [animate_button(button_limpiar), clear_fields()], style="Button.TButton")
button_limpiar.grid(row=len(labels) + 1, column=1, pady=(5, 0), padx=(5, 10), sticky='we')

success_label = ttk.Label(frame, text="", font=common_font)
success_label.grid(row=2, column=0, columnspan=3, sticky="we", pady=(0, 10))

version_label = ttk.Label(frame, text=f"Versión {version}", font=common_font)
version_label.grid(row=3, column=2, sticky='se', padx=(0, 10), pady=(10, 0))

for i in range(4):
    window.grid_columnconfigure(i, weight=1)
    window.grid_rowconfigure(i, weight=1)

window.mainloop()
